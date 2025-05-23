import os
import logging
import sys

# 全局变量，用于存储 openpyxl 模块
openpyxl_available = False
try:
    import openpyxl
    openpyxl_available = True
except ImportError:
    error_msg = "警告: 未找到 openpyxl 模块，Excel 功能将无法正常工作"
    print(error_msg)
    logging.warning(error_msg)

# 导入 pandas
try:
    import pandas as pd
    # 不要全局设置引擎，而是在每次调用时指定
except ImportError as e:
    print(f"警告: {str(e)}")
    logging.warning(f"导入 pandas 失败: {str(e)}")

class ExcelFileNotFound(Exception):
    pass

def get_excel_info(import_root_dir, person_name, person_id, class_code):
    """
    从Excel获取文件相关信息。
    :param import_root_dir: Excel文件根目录
    :param person_name: 人名
    :param person_id: 编号
    :param class_code: 类号
    :return: (material_name, file_date, page_count, data_list)
    :raises ExcelFileNotFound: 未找到匹配的Excel文件
    """
    material_name = ""
    file_date = ""
    page_count = ""
    data_list = []
    
    # 记录调试信息
    logging.info("="*50)
    logging.info(f"开始处理Excel信息")
    logging.info(f"- 根目录: {os.path.abspath(import_root_dir)}")
    logging.info(f"- 搜索条件: 姓名='{person_name}', 编号='{person_id}', 类号='{class_code}'")

    # 解析sheet名
    class_parts = class_code.split('-')
    main_code = class_parts[0]
    sheet_code = main_code
    if main_code in ['4', '9'] and len(class_parts) >= 2:
        sheet_code = f"{main_code}-{class_parts[1]}"
    sheet_map = {
        '1': '一', '2': '二', '3': '三',
        '4-1': '四-1', '4-2': '四-2', '4-3': '四-3', '4-4': '四-4',
        '5': '五', '6': '六', '7': '七', '8': '八',
        '9-1': '九-1', '9-2': '九-2', '9-3': '九-3', '9-4': '九-4',
        '10': '十'
    }
    sheet_name = sheet_map.get(sheet_code)
    if not sheet_name:
        return material_name, file_date, page_count, data_list

    # 只搜索'目录'子目录
    catalog_dir = os.path.join(import_root_dir, '目录')
    if not os.path.exists(catalog_dir):
        logging.warning(f"目录不存在: {catalog_dir}")
        raise ExcelFileNotFound(f"目录不存在: {catalog_dir}")
    
    logging.info(f"正在搜索目录: {catalog_dir}")
    
    # 构建预期的文件名（编号+姓名，无空格）
    expected_name = f"{person_id}{person_name}"
    expected_filename = f"{expected_name}.xlsx"
    logging.info(f"正在搜索Excel文件，目标文件名: {expected_filename}")
    file_path = os.path.join(catalog_dir, expected_filename)
    
    # 检查文件是否存在
    if os.path.exists(file_path):
        abs_path = os.path.abspath(file_path)
        logging.info(f"找到匹配的Excel文件: {abs_path}")
        logging.info(f"正在读取工作表: {sheet_name}")
        try:
            material_name, file_date, page_count, data_list = read_excel_file(file_path, sheet_name, class_code)
            logging.info(f"成功读取Excel文件: {os.path.basename(file_path)}, 工作表: {sheet_name}")
            return material_name, file_date, page_count, data_list
        except Exception as e:
            logging.error(f"读取Excel文件失败: {str(e)}")
            raise
    
    # 如果没有找到匹配的文件，记录日志并抛出异常
    error_msg = f"未找到匹配的Excel文件: {expected_filename} (在目录: {catalog_dir})"
    logging.error(error_msg)
    logging.error(f"请确保文件按照'编号+姓名.xlsx'的格式命名，例如: {expected_filename}")
    raise ExcelFileNotFound(error_msg)

def read_excel_file(file_path, sheet_name, class_code=None, file_index=None):
    """
    读取Excel文件并返回数据
    :param file_path: Excel文件路径
    :param sheet_name: 工作表名称
    :param class_code: 类号（例如：'1', '4-1', '9-2'等）
    :param file_index: 文件索引（例如：'1-1'中的1-1）
    :return: (material_name, file_date, page_count, data_list)
              material_name: 材料名称
              file_date: 文件日期
              page_count: 页数
              data_list: 包含所有行的数据列表，每行格式为 [类号, 材料名称, 日期, 页数]
    """
    material_name = ""
    file_date = ""
    page_count = ""
    data_list = []
    
    logging.info(f"开始读取Excel文件: {os.path.abspath(file_path)}")
    logging.info(f"工作表名称: {sheet_name}, 类号: {class_code}, 文件索引: {file_index}")
    
    # 检查 openpyxl 是否可用
    if not openpyxl_available:
        error_msg = "错误: 未找到 openpyxl 模块，无法读取 Excel 文件。请安装: pip install openpyxl"
        logging.error(error_msg)
        print(error_msg)
        return "", "", "", []
    
    # 根据类号确定起始行
    start_row = 5  # 默认从第5行开始
    if class_code:
        main_code = class_code.split('-')[0]
        if main_code in ['4', '9']:  # 第四和第九类从第6行开始
            start_row = 6
            logging.info(f"检测到第{main_code}类，从第{start_row}行开始读取")
        else:  # 其他类从第5行开始
            logging.info(f"检测到第{main_code}类，从第{start_row}行开始读取")
    else:
        logging.warning(f"未指定类号，使用默认起始行: {start_row}")
    
    if not os.path.exists(file_path):
        logging.error(f"文件不存在: {file_path}")
        return material_name, file_date, page_count, data_list
    
    wb = None
    try:
        # 使用openpyxl读取Excel文件
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        # 检查工作表是否存在
        if sheet_name not in wb.sheetnames:
            logging.warning(f"工作表 '{sheet_name}' 不存在，使用第一个可用工作表")
            sheet_name = wb.sheetnames[0]
            
        ws = wb[sheet_name]
        
        # 使用固定的列索引
        # A列(索引0): 类号
        # B列(索引1): 材料名称
        # C/D/E列(索引2/3/4): 日期（年/月/日）
        # F列(索引5): 页数
        name_col = 1  # B列
        date_cols = [2, 3, 4]  # C/D/E列
        page_col = 5  # F列
        
        logging.info("使用固定列索引:")
        logging.info(f"- 类号: A列(索引0)")
        logging.info(f"- 材料名称: B列(索引{name_col})")
        logging.info(f"- 日期: C/D/E列(索引{date_cols}): 年/月/日")
        logging.info(f"- 页数: F列(索引{page_col})")
        logging.info(f"- 起始行: 第{start_row}行")
        
        # 从起始行开始读取数据
        row_count = 0
        empty_row_count = 0
        max_empty_rows = 7  # 连续空行的最大数量
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            # 检查是否连续多行为空
            if not any(cell is not None and str(cell).strip() for cell in row):
                empty_row_count += 1
                if empty_row_count >= max_empty_rows:
                    logging.info(f"检测到连续 {empty_row_count} 个空行，停止读取")
                    break
                continue
            else:
                empty_row_count = 0
            
            # 获取类号
            row_class = str(row[0]).strip() if row[0] is not None else ""
            
            # 获取材料名称
            material = str(row[name_col]).strip() if len(row) > name_col and row[name_col] is not None else ""
            
            # 获取日期（年/月/日）
            year = str(int(row[date_cols[0]])) if len(row) > date_cols[0] and row[date_cols[0]] is not None and str(row[date_cols[0]]).strip() != '' else ""
            month = str(int(row[date_cols[1]])) if len(date_cols) > 1 and len(row) > date_cols[1] and row[date_cols[1]] is not None and str(row[date_cols[1]]).strip() != '' else ""
            day = str(int(row[date_cols[2]])) if len(date_cols) > 2 and len(row) > date_cols[2] and row[date_cols[2]] is not None and str(row[date_cols[2]]).strip() != '' else ""
            
            # 组合日期
            date_parts = [year, month, day]
            date_str = '-'.join([d for d in date_parts if d])
            
            # 获取页数
            pages = str(int(row[page_col])) if len(row) > page_col and row[page_col] is not None and str(row[page_col]).strip() != '' else "0"
            
            # 添加到数据列表
            data_list.append([row_class, material, date_str, pages])
            row_count += 1
        
        logging.info(f"读取完成 - 共 {row_count} 行数据")
        
        # 如果传入了file_index，则根据文件索引匹配行
        if file_index and data_list:
            try:
                # 计算行索引（从0开始）
                # 例如：file_index='1-1' -> row_idx = 0 (第1个文件对应第1行数据)
                # file_index='1-2' -> row_idx = 1 (第2个文件对应第2行数据)
                file_num = int(file_index.split('-')[-1])  # 获取文件序号
                row_idx = file_num - 1  # 转换为0-based索引
                
                if 0 <= row_idx < len(data_list):
                    material_name = data_list[row_idx][1]  # 材料名称
                    file_date = data_list[row_idx][2]      # 日期
                    page_count = data_list[row_idx][3]     # 页数
                    logging.info(f"根据文件索引 '{file_index}' 获取到数据: 材料名称='{material_name}', 日期='{file_date}', 页数='{page_count}'")
                else:
                    logging.warning(f"文件索引 '{file_index}' 超出范围，有效范围: 1-{len(data_list)}")
            except (ValueError, IndexError) as e:
                logging.warning(f"无效的文件索引格式: {file_index}, 错误: {str(e)}")
        # 如果传入了class_code，则尝试匹配对应的行
        elif class_code and data_list:
            # 移除class_code中的文件扩展名（如果有）
            class_code_base = class_code.split('.')[0]
            
            # 在数据列表中查找匹配的类号
            for item in data_list:
                if item[0] == class_code_base:  # item[0]是类号
                    material_name = item[1]  # 材料名称
                    file_date = item[2]      # 日期
                    page_count = item[3]     # 页数
                    logging.info(f"找到匹配的类号 '{class_code_base}': 材料名称='{material_name}', 日期='{file_date}', 页数='{page_count}'")
                    break
            else:
                logging.warning(f"未找到匹配的类号: {class_code_base}")
            
            return material_name, file_date, page_count, data_list
        
        logging.info(f"读取完成 - 共 {row_count} 行数据，其中 {len(data_list)} 行非空")
        
        # 记录提取的信息
        logging.info("="*50)
        logging.info(f"提取的信息:")
        logging.info(f"- 材料名称: '{material_name}'")
        logging.info(f"- 日期: '{file_date}'")
        logging.info(f"- 页数: '{page_count}'")
        logging.info(f"- 数据行数: {len(data_list)}")
        # 记录第一行数据
        if data_list and len(data_list) > 0:
            logging.info("第一行数据示例:")
            for idx, value in enumerate(data_list[0]):
                logging.info(f"  列{idx}: {value}")
        logging.info("="*50)
        
    except Exception as e:
        logging.error(f"读取Excel文件失败: {str(e)}", exc_info=True)
        logging.error(f"错误发生在处理文件: {file_path}")
        logging.error(f"工作表: {sheet_name}")
    finally:
        # 确保工作簿被正确关闭
        if wb is not None:
            wb.close()
    
    return material_name, file_date, page_count, data_list
